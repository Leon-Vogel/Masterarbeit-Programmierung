from _dir_init import *
from azure.storage.blob import BlobClient, ContainerClient, BlobServiceClient  # pip install azure-storage-blob
from azure.identity import DefaultAzureCredential  # pip install azure-identity
import openpyxl
import io
import traceback
import pandas as pd
"""
WICHTIG!

Damit der Zugriff auf die DataContainer im Storage-Account funktioniert,
muss der Zugreifende Computer zunächst authentifiziert werden.

Siehe dazu Beschreibun in der _raise_exc Methode

"""


class MieleDataContainer:
    GTG = "gtg"
    SMART_HOME = "smarthome"


class MieleStorageAccount:
    DEV = "supwedevstadata"


def _raise_exc():
    print("\nWICHTIG:")
    print("===========================================================================")
    print("Es wird versucht, auf die Inhalte eines DataContainers zuzugreifen.")
    print(
        "Entweder ist (A) der aktuelle Computer nicht eingerichtet oder (B) die IP-Adresse ist nicht in der Cloud eingetragen."
    )
    print("---")
    print("Bzgl. (A) werden DefaultCredentials verwendet, die erst auf diesem Computer hinterlegt werden müssen.")
    print(
        "Dazu muss man zunächst die AZURE CLI installieren und sich dann über PowerShell/CommandShell in Azure einloggen."
    )
    print("Befehl: az login")
    print("---")
    print("Bzgl. (B): Es muss die aktuelle IP in den Azure Netzwerkeinstellungen hinterlegt sein.")
    print("Gehe zum Storage-Account, klicke im Menü auf Networking und Hinterlege die IP in der Liste.")
    print("===========================================================================")
    print("\n\nORIGINAL EXCEPTION:")
    traceback.print_exc()


def decode_textfile(rawcontent, encoding):
    return rawcontent.decode(encoding)


def decode_excelfile(rawcontent):
    byte_stream = io.BytesIO(rawcontent)
    workbook = openpyxl.load_workbook(byte_stream)
    return workbook


def _get_container_client(container, storage_acc):
    credential = DefaultAzureCredential()
    acc_url = f"https://{storage_acc}.blob.core.windows.net"
    blob_service_client = BlobServiceClient(account_url=acc_url, credential=credential)
    container_client = blob_service_client.get_container_client(container)
    return container_client


def get_bloblist_from_container(container=MieleDataContainer.GTG, storage_account=MieleStorageAccount.DEV):
    """list all files"""
    try:
        client = _get_container_client(container, storage_account)
        blobs = client.list_blobs()
        return list(blobs)
    except:
        _raise_exc()


def _get_content_from_container(container, stor_acc, files):
    try:
        cont_client = _get_container_client(container, stor_acc)
        contents = {}
        for file in files:
            blob_client = cont_client.get_blob_client(file)
            content = blob_client.download_blob().readall()
            contents[file] = content
        return contents
    except:
        _raise_exc()


def get_binaryfiles_from_container(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    blobs_filenames=[
        "dummy.txt",
    ],
):
    """lade Binärdateien, die über die Komplexität von Textdateien hinausgehen.
    Z.B. EXE-Dateien, Office-Dateien usw."""
    return _get_content_from_container(container, storage_account, blobs_filenames)

def get_binaryfile_single_from_container(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    blob_filename="dummy.txt",
):
    """lade eine einzelne Binärdatei, die über die Komplexität von Textdateien hinausgeht.
    Z.B. EXE-Dateien, Office-Dateien usw."""
    return _get_content_from_container(container, storage_account, [blob_filename])[blob_filename]


def get_textfiles_from_container(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    blobs_filenames_with_encoding=[
        ("dummy.txt", "utf-8"),
    ],
):
    """lade Textdateien in einem definierten Format, z.B: UTF-8 oder ASCII"""
    files = [file for file, encoding in blobs_filenames_with_encoding]
    contents_downloaded = _get_content_from_container(container, storage_account, files)
    for file, enconding in blobs_filenames_with_encoding:
        if enconding == "":
            enconding = "utf-8"
        content = contents_downloaded[file]
        content_decoded = decode_textfile(content, enconding)
        contents_downloaded[file] = content_decoded
    return contents_downloaded


def get_all_files_from_directory_recursive(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    directory="simulation_data/flexsim_dummy/",
):
    client = _get_container_client(container, storage_account)
    blobs = client.list_blobs(name_starts_with=directory)
    blobs_filenames = [blob.name for blob in blobs]
    return _get_content_from_container(container, storage_account, blobs_filenames)


def get_excelfiles_from_container(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    excel_filenames=[
        "dummy.xlsx",
    ],
):
    contents_downloaded = get_binaryfiles_from_container(container, storage_account, excel_filenames)
    for file, content in contents_downloaded.items():
        byte_stream = io.BytesIO(content)
        workbook = openpyxl.load_workbook(byte_stream)
        contents_downloaded[file] = workbook
    return contents_downloaded


def get_excelfile_single_from_container(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    excel_filename="dummy.xlsx",
    sheet_name=None,
):
    content_downloaded = get_excelfiles_from_container(container, storage_account, [excel_filename])
    workbook = content_downloaded[excel_filename]
    if sheet_name == None:
        return workbook
    sheet = workbook[sheet_name]
    # load table in sheet
    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = []
        for value in row:
            row_data.append(value)
        sheet_data.append(row_data)
    return sheet_data




def upload_file(
    container=MieleDataContainer.GTG,
    storage_account=MieleStorageAccount.DEV,
    file_path_target="simulation_data/_testfile_can_be_deleted.txt",
    content_encoded="ASPIDJ23",
):
    cont_cline = _get_container_client(container, storage_account)
    cont_cline.upload_blob(name=file_path_target, data=content_encoded)


if __name__ == "__main__":
    upload_file()
    testdir = get_all_files_from_directory_recursive()
    bloblist = get_bloblist_from_container()

    testcontent = get_excelfiles_from_container(excel_filenames=["simulation_data/flexsim_dummy/Endprog_Woche1.xlsx"])

    pass
